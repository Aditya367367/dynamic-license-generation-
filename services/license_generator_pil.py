#!/usr/bin/env python3
"""
CINEFIL Performance License Generator using PIL (Pillow)
Works with JPG template for precise text placement on your actual license form.
"""

import os
import json
import mimetypes
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime
import qrcode
from PIL import Image, ImageDraw, ImageFont
import re
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

# Load environment variables from .env file
load_dotenv()

class LicenseGeneratorPIL:
    def __init__(self, config_path: str = "field_mapping.json"):
        """Initialize the license generator with field mapping configuration."""
        self.config_path = config_path
        self.field_mapping = self._load_field_mapping()
        
        # Color constants
        self.COLORS = {
            'title': (201, 166, 70),    # #C9A646 (gold)
            'labels': (0, 0, 0),         # #000000 (black)
            'data': (17, 17, 17),        # #111111 (dark gray)
            'footer': (0, 0, 0)          # #000000 (black)
        }

        self.FIELD_LAYOUTS = {
            "Licence Number": {
                "font_scale": 0.64,
                "x_shift": 12,
                "y_shift": -16,
                "padding_y_ratio": 0.10
            },
            "Validity From": {
                "font_scale": 0.58,
                "y_shift": -16,
                "padding_y_ratio": 0.10
            },
            "Name of the Licensee": {
                "multiline": True,
                "max_lines": 2,
                "font_scale": 0.74,
                "y_shift": -28,
                "padding_y_ratio": 0.05
            },
            "Type of Premise": {
                "font_scale": 0.66,
                "y_shift": -72,
                "padding_y_ratio": 0.08
            },
            "License Category": {
                "font_scale": 0.66,
                "y_shift": -78,
                "padding_y_ratio": 0.08
            },
            "Address of Premise": {
                "multiline": True,
                "font_scale": 0.64,
                "y_shift": -78,
                "padding_y_ratio": 0.03
            }
        }
        
    def _load_field_mapping(self) -> dict:
        """Load field mapping from JSON configuration."""
        try:
            with open(self.config_path, 'r') as f:
                config = json.load(f)
                return {field['field_name']: field for field in config['fields']}
        except FileNotFoundError:
            raise
        except json.JSONDecodeError:
            raise

    def _sanitize_filename(self, text: str) -> str:
        """Sanitize text for use in filename."""
        sanitized = re.sub(r'[^\w\s-]', '', text)
        sanitized = re.sub(r'[-\s]+', '_', sanitized)
        return sanitized.strip('_')[:50]

    def update_license_register(self, user_data: dict, result_data: dict,
                                register_path: str = "license_register.xlsx") -> str:
        """Append generated license details to a real Excel register."""

        register_fields = [
            "Generated At",
            "Licence Number",
            "Validity From",
            "Name of the Licensee",
            "Type of Premise",
            "License Category",
            "Address of Premise",
            "QR Code",
            "Link"
        ]

        row = {
            "Generated At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Licence Number": user_data.get("Licence Number", ""),
            "Validity From": user_data.get("Validity From", ""),
            "Name of the Licensee": user_data.get("Name of the Licensee", ""),
            "Type of Premise": user_data.get("Type of Premise", ""),
            "License Category": user_data.get("License Category", ""),
            "Address of Premise": user_data.get("Address of Premise", ""),
            "QR Code": "",
            "Link": result_data.get("qr_target_link", ""),
        }
        qr_code_path = os.path.abspath(result_data.get("qr_code", ""))

        if os.path.exists(register_path):
            workbook = load_workbook(register_path)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Licenses"
            worksheet.append(register_fields)

            for index, header in enumerate(register_fields, start=1):
                cell = worksheet.cell(row=1, column=index)
                cell.font = Font(bold=True)

        worksheet.append([row[field] for field in register_fields])
        last_row = worksheet.max_row

        link_columns = {"Link"}
        for index, header in enumerate(register_fields, start=1):
            if header in link_columns:
                cell = worksheet.cell(row=last_row, column=index)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        qr_column_index = register_fields.index("QR Code") + 1
        qr_cell = worksheet.cell(row=last_row, column=qr_column_index)
        worksheet.row_dimensions[last_row].height = 90
        if qr_code_path and os.path.exists(qr_code_path):
            qr_image = XLImage(qr_code_path)
            qr_image.width = 84
            qr_image.height = 84
            worksheet.add_image(qr_image, qr_cell.coordinate)

        column_widths = {
            "Generated At": 22,
            "Licence Number": 22,
            "Validity From": 16,
            "Name of the Licensee": 30,
            "Type of Premise": 24,
            "License Category": 24,
            "Address of Premise": 40,
            "QR Code": 16,
            "Link": 38
        }

        for index, header in enumerate(register_fields, start=1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = column_widths.get(header, 24)

        workbook.save(register_path)
        return register_path

    def get_user_input(self) -> dict:
        """Collect user input for all required fields."""
        
        user_data = {}
        
        for field_name, field_config in self.field_mapping.items():
            # Skip QR Code field - no user input needed
            if field_name == "QR Code":
                continue
                
            if field_config['type'] == 'date':
                default_date = datetime.now().strftime("%d-%m-%Y")
                user_input = input(f"{field_name} (format: DD-MM-YYYY) [{default_date}]: ").strip()
                user_data[field_name] = user_input if user_input else default_date
            else:
                user_input = input(f"{field_name}: ").strip()
                user_data[field_name] = user_input
                
        return user_data

    def _get_font(self, font_type: str = "data", size: int = 28) -> ImageFont.FreeTypeFont:
        """Get specific font for license text rendering.

        Ubuntu/Fedora hosts can ship different font packages. We therefore try a
        broader set of distro-specific absolute paths and generic font family
        names before falling back to PIL's tiny bitmap default font.
        """
        font_candidates = {
            "title": [
                "/usr/share/fonts/julietaula-montserrat-fonts/Montserrat-Bold.otf",
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Bold.ttf",
                "/usr/share/fonts/open-sans/OpenSans-Semibold.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "/usr/share/fonts/truetype/liberation2/LiberationSerif-Bold.ttf",
                "DejaVuSans-Bold.ttf",
            ],
            "label": [
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Bold.ttf",
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Regular.ttf",
                "/usr/share/fonts/open-sans/OpenSans-Semibold.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
                "/usr/share/fonts/truetype/liberation2/LiberationSerif-Bold.ttf",
                "DejaVuSerif-Bold.ttf",
            ],
            "script": [
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Italic.ttf",
                "/usr/share/fonts/google-noto-vf/NotoSerif-Italic[wght].ttf",
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Regular.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Italic.ttf",
                "/usr/share/fonts/truetype/liberation2/LiberationSerif-Italic.ttf",
                "DejaVuSerif-Italic.ttf",
                "DejaVuSerif.ttf",
            ],
            "data": [
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Regular.ttf",
                "/usr/share/fonts/open-sans/OpenSans-Regular.ttf",
                "/usr/share/fonts/google-noto-vf/NotoSerif[wght].ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
                "/usr/share/fonts/truetype/liberation2/LiberationSerif-Regular.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "DejaVuSerif.ttf",
                "DejaVuSans.ttf",
            ],
            "small": [
                "/usr/share/fonts/open-sans/OpenSans-Regular.ttf",
                "/usr/share/fonts/google-crosextra-caladea-fonts/Caladea-Regular.ttf",
                "/usr/share/fonts/google-droid-sans-fonts/DroidSans.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "DejaVuSans.ttf",
            ],
        }

        candidates = font_candidates.get(font_type, font_candidates["data"])
        for font_candidate in candidates:
            # Absolute paths should exist; family-name candidates are resolved by PIL.
            if font_candidate.startswith("/") and not Path(font_candidate).exists():
                continue
            try:
                return ImageFont.truetype(font_candidate, size)
            except OSError:
                continue

        # Final fallback (keeps pipeline alive, but should now be rare).
        return ImageFont.load_default()

    def draw_spaced_text(self, draw: ImageDraw, position: tuple, text: str, font: ImageFont.FreeTypeFont, 
                        fill: tuple = (0, 0, 0), spacing: int = 2) -> None:
        """Draw text with letter spacing for field labels."""
        x, y = position
        for char in text:
            draw.text((x, y), char, font=font, fill=fill)
            x += draw.textlength(char, font=font) + spacing

    def _get_scaled_font(self, field_name: str, template_width: int = 0) -> ImageFont.FreeTypeFont:
        """Return a starting font size before field-level fitting is applied."""
        base_sizes = {
            "Licence Number": 120,
            "Validity From": 100,
            "Name of the Licensee": 140,
            "Type of Premise": 120,
            "License Category": 120,
            "Address of Premise": 110
        }

        base_size = base_sizes.get(field_name, 70)
        multiplier = template_width / 2480.0 if template_width else 1.0
        
        # Add system-specific font scaling to handle Ubuntu vs Fedora differences
        import platform
        system_scale = 1.0
        if platform.system() == "Linux":
            # Check if we're on Ubuntu (which typically has smaller font rendering)
            try:
                with open('/etc/os-release', 'r') as f:
                    os_info = f.read().lower()
                    if 'ubuntu' in os_info:
                        system_scale = 1.25  # Increase font size for Ubuntu
                    elif 'fedora' in os_info:
                        system_scale = 1.0  # Keep original size for Fedora
            except:
                # Fallback: if we can't detect OS, use a moderate scale
                system_scale = 1.15
        
        final_size = max(24, int(base_size * multiplier * system_scale))
        return self._get_font("script", final_size)

    def _get_text_size(self, draw: ImageDraw, text: str, font: ImageFont.FreeTypeFont) -> tuple[int, int]:
        """Measure rendered text size."""
        bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0], bbox[3] - bbox[1]

    def _wrap_text_to_width(self, draw: ImageDraw, text: str, max_width: int,
                            font: ImageFont.FreeTypeFont) -> list[str]:
        """Wrap text into multiple lines that fit within the given width."""
        words = text.split()
        if not words:
            return [""]

        lines = []
        current_line = ""

        for word in words:
            test_line = f"{current_line} {word}".strip()
            text_width, _ = self._get_text_size(draw, test_line, font)

            if text_width <= max_width or not current_line:
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word

        if current_line:
            lines.append(current_line)

        return lines

    def _fit_text_font(self, draw: ImageDraw, text: str, field_name: str, max_width: int,
                       max_height: int, multiline: bool = False, max_lines: int | None = None,
                       font_scale: float = 1.0) -> tuple[ImageFont.FreeTypeFont, list[str], int]:
        """Fit text to a field box by shrinking from a field-specific starting size."""
        start_font = self._get_scaled_font(field_name)
        start_size = max(24, int(getattr(start_font, "size", 70) * font_scale))

        for size in range(start_size, 23, -2):
            font = self._get_font("script", size)
            if multiline:
                lines = self._wrap_text_to_width(draw, text, max_width, font)
                if max_lines is not None and len(lines) > max_lines:
                    continue
                _, line_height = self._get_text_size(draw, "Ag", font)
                line_gap = max(8, int(line_height * 0.25))
                total_height = len(lines) * line_height + max(0, len(lines) - 1) * line_gap
                widest_line = max((self._get_text_size(draw, line, font)[0] for line in lines), default=0)

                if widest_line <= max_width and total_height <= max_height:
                    return font, lines, line_height + line_gap
            else:
                text_width, text_height = self._get_text_size(draw, text, font)
                if text_width <= max_width and text_height <= max_height:
                    return font, [text], 0

        fallback_font = self._get_font("script", 24)
        if multiline:
            lines = self._wrap_text_to_width(draw, text, max_width, fallback_font)
            if max_lines is not None:
                lines = lines[:max_lines]
            _, line_height = self._get_text_size(draw, "Ag", fallback_font)
            line_gap = max(6, int(line_height * 0.2))
            return fallback_font, lines, line_height + line_gap
        return fallback_font, [text], 0

    def draw_centered_text(self, draw: ImageDraw, text: str, x: int, y: int, w: int, h: int, 
                         font: ImageFont.FreeTypeFont, fill: tuple) -> None:
        """Draw text centered within a rectangle."""
        bbox = draw.textbbox((0, 0), text, font=font)
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
        new_x = x + (w - text_w) // 2
        new_y = y + (h - text_h) // 2
        draw.text((new_x, new_y), text, fill=fill, font=font)

    def draw_bold_text(self, draw: ImageDraw, position: tuple, text: str, font: ImageFont.FreeTypeFont, 
                      fill: tuple) -> None:
        """Draw text with slight bold effect for professional look."""
        x, y = position
        draw.text((x, y), text, font=font, fill=fill)
        draw.text((x+1, y), text, font=font, fill=fill)

    def _draw_wrapped_lines(self, draw: ImageDraw, lines: list[str], x: int, y: int,
                            font: ImageFont.FreeTypeFont, fill: tuple = (0, 0, 0),
                            line_height: int = 30) -> None:
        """Draw already wrapped text lines using the fitted line spacing."""
        for i, line in enumerate(lines):
            draw.text((x, y + (i * line_height)), line, fill=fill, font=font)

    def fill_jpg_template(self, template_path: str, user_data: dict, output_path: str) -> str:
        """Fill JPG template with user data using pixel-perfect positioning and scaled fonts."""
        
        # Load the JPG template
        try:
            image = Image.open(template_path)
            img_width, img_height = image.size
            draw = ImageDraw.Draw(image)
        except Exception as e:
            raise
        
        # Add text for each field with proper scaling and centering
        for field_name, field_config in self.field_mapping.items():
            # Skip QR field as it's handled separately
            if field_name == "QR Code":
                continue
                
            if field_name in user_data:
                value = user_data[field_name]
                x = field_config['x']
                y = field_config['y']
                w = field_config['width']
                h = field_config['height']

                layout = self.FIELD_LAYOUTS.get(field_name, {})
                render_as_multiline = layout.get("multiline", field_config['type'] == 'multiline')
                font_scale = layout.get("font_scale", 1.0)
                x_shift = layout.get("x_shift", 0)
                y_shift = layout.get("y_shift", 0)
                max_lines = layout.get("max_lines")

                if render_as_multiline:
                    padding_x = max(20, int(w * layout.get("padding_x_ratio", 0.03)))
                    padding_y = max(12, int(h * layout.get("padding_y_ratio", 0.08)))
                    available_width = w - (padding_x * 2)
                    available_height = h - (padding_y * 2)
                    font, lines, line_height = self._fit_text_font(
                        draw,
                        value,
                        field_name,
                        available_width,
                        available_height,
                        multiline=True,
                        max_lines=max_lines,
                        font_scale=font_scale
                    )

                    self._draw_wrapped_lines(
                        draw,
                        lines,
                        x + padding_x + x_shift,
                        y + padding_y + y_shift,
                        font,
                        fill=self.COLORS['data'],
                        line_height=line_height
                    )
                else:
                    padding_x = max(20, int(w * layout.get("padding_x_ratio", 0.04)))
                    padding_y = max(8, int(h * layout.get("padding_y_ratio", 0.12)))
                    available_width = w - (padding_x * 2)
                    available_height = h - (padding_y * 2)
                    font, _, _ = self._fit_text_font(
                        draw,
                        value,
                        field_name,
                        available_width,
                        available_height,
                        multiline=False,
                        font_scale=font_scale
                    )

                    _, text_height = self._get_text_size(draw, value, font)
                    text_x = x + padding_x + x_shift
                    text_y = y + max(padding_y, (h - text_height) // 2) + y_shift
                    draw.text((text_x, text_y), value, fill=self.COLORS['data'], font=font)
        
        # Save the filled image
        image.save(output_path, 'JPEG', quality=95)
        return output_path

    def convert_to_pdf(self, jpg_path: str, pdf_path: str) -> str:
        """Convert JPG to PDF for compatibility with existing pipeline."""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            
            # Get image dimensions
            img = Image.open(jpg_path)
            img_width, img_height = img.size
            
            # Create PDF with same dimensions as image
            c = canvas.Canvas(pdf_path, pagesize=(img_width, img_height))
            c.drawImage(jpg_path, 0, 0, img_width, img_height)
            c.save()
            
            return pdf_path
        except ImportError:
            return jpg_path

    def compress_image(self, input_path: str, output_path: str) -> str:
        """Compress image for upload."""
        
        try:
            img = Image.open(input_path)

            # Downscale the upload copy to keep the WordPress PDF lightweight.
            max_width = 1800
            if img.width > max_width:
                new_height = int((max_width / img.width) * img.height)
                img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)

            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")

            # Save with stronger compression for the upload-only derivative.
            img.save(output_path, 'JPEG', quality=38, optimize=True, progressive=True)
            return output_path
        except Exception as e:
            return input_path

    def compress_pdf(self, input_pdf: str, output_pdf: str) -> str:
        """Compress PDF for upload using Ghostscript."""
        
        try:
            # Use Ghostscript for PDF compression
            import subprocess
            import tempfile
            
            # Always use a temporary file to avoid same-file issues
            temp_output = f"{output_pdf}.tmp"
            
            cmd = (
                f'gs -sDEVICE=pdfwrite '
                f'-dCompatibilityLevel=1.4 '
                f'-dPDFSETTINGS=/screen '
                f'-dDetectDuplicateImages=true '
                f'-dDownsampleColorImages=true '
                f'-dColorImageDownsampleType=/Bicubic '
                f'-dColorImageResolution=96 '
                f'-dDownsampleGrayImages=true '
                f'-dGrayImageDownsampleType=/Bicubic '
                f'-dGrayImageResolution=96 '
                f'-dDownsampleMonoImages=true '
                f'-dMonoImageDownsampleType=/Subsample '
                f'-dMonoImageResolution=150 '
                f'-dNOPAUSE -dQUIET -dBATCH '
                f'-sOutputFile="{temp_output}" "{input_pdf}"'
            )
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
            
            if result.returncode == 0 and os.path.exists(temp_output):
                # Replace the original file with the compressed version
                os.replace(temp_output, output_pdf)
                return output_pdf
            else:
                # Clean up temp file if it exists
                if os.path.exists(temp_output):
                    os.remove(temp_output)
                # Fallback: just return the original file (no compression)
                return input_pdf
                
        except Exception as e:
            # Clean up temp file if it exists
            if 'temp_output' in locals() and os.path.exists(temp_output):
                os.remove(temp_output)
            # Fallback: just return the original file (no compression)
            return input_pdf

    def upload_to_wordpress(self, file_path: str, wp_url: str, username: str, app_password: str,
                            upload_filename: str | None = None) -> str:
        """Upload only PDF files to WordPress using REST API."""

        if not file_path.lower().endswith(".pdf"):
            raise ValueError(f"Only PDF uploads are allowed. Received: {file_path}")
        
        if not wp_url.endswith('/'):
            wp_url += '/'
        
        upload_url = f"{wp_url}wp-json/wp/v2/media"
        upload_name = upload_filename or os.path.basename(file_path)
        mime_type = mimetypes.guess_type(upload_name)[0] or "application/octet-stream"
        
        try:
            with open(file_path, 'rb') as f:
                files = {'file': (upload_name, f, mime_type)}
                headers = {'Content-Disposition': f'attachment; filename="{upload_name}"'}
                
                response = requests.post(
                    upload_url,
                    files=files,
                    headers=headers,
                    auth=HTTPBasicAuth(username, app_password),
                    timeout=30
                )
            
            response.raise_for_status()
            result = response.json()
            source_url = result['source_url']
            
            return source_url
            
        except requests.exceptions.RequestException as e:
            raise

    def generate_qr_code(self, data: str, output_path: str, size: int = 200) -> str:
        """Generate QR code from the provided data."""
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(output_path)
        
        return output_path

    def embed_qr_in_image(self, base_image_path: str, qr_path: str, output_path: str) -> str:
        """Embed QR code into base image using field mapping coordinates."""
        
        try:
            # Open images
            base_img = Image.open(base_image_path)
            qr_img = Image.open(qr_path)
            
            # Get QR field coordinates from field mapping
            qr_field = self.field_mapping.get("QR Code")
            if qr_field:
                qr_x = qr_field['x']
                qr_y = qr_field['y'] + 10  # Position 10px below field
                qr_width = qr_field['width']
                qr_height = qr_field['height']
                
                # Resize QR to fit field dimensions
                qr_img = qr_img.resize((qr_width, qr_height), Image.Resampling.LANCZOS)
            else:
                # Fallback: bottom-right corner
                img_width, img_height = base_img.size
                qr_size = 200
                qr_x = img_width - qr_size - 20
                qr_y = img_height - qr_size - 20
                qr_img = qr_img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
            
            # Paste QR onto base image
            base_img.paste(qr_img, (qr_x, qr_y))
            base_img.save(output_path, 'JPEG', quality=95)
            
            return output_path
        except Exception as e:
            raise

    def run_complete_pipeline(self, template_path: str, wp_url: str, username: str, app_password: str, user_data: dict = None, register_path: str = None, output_dir: str = None):
        """Run the complete license generation pipeline using PIL."""
        try:
            # Get environment mode
            environment = os.getenv('ENVIRONMENT', 'production').lower()
            is_test_mode = environment == 'test'
            
            if is_test_mode:
                pass
            
            # Step 1: Get user input
            if user_data is None:
                user_data = self.get_user_input()
            
            # Generate dynamic filenames and folder
            license_number = self._sanitize_filename(user_data.get('Licence Number', 'UNKNOWN'))
            licensee_name = self._sanitize_filename(user_data.get('Name of the Licensee', 'UNKNOWN'))
            base_filename = f"{license_number}_{licensee_name}"
            
            # Use provided output directory or current directory
            folder_name = f"{license_number}_{licensee_name}"
            if output_dir:
                folder_path = output_dir
            else:
                os.makedirs(folder_name, exist_ok=True)
                folder_path = folder_name
            
            # File paths
            filled_jpg = os.path.join(folder_path, f"{base_filename}_filled.jpg")
            filled_pdf = os.path.join(folder_path, f"{base_filename}_filled.pdf")
            compressed_jpg = os.path.join(folder_path, f"{base_filename}_compressed.jpg")
            compressed_pdf = os.path.join(folder_path, f"{base_filename}_compressed.pdf")
            qr_path = os.path.join(folder_path, f"{base_filename}_qr.png")
            final_jpg = os.path.join(folder_path, f"{base_filename}_final.jpg")
            final_pdf = os.path.join(folder_path, f"{base_filename}_final.pdf")
            wordpress_upload_name = f"{base_filename}_license.pdf"
            
            # Step 2: Fill JPG template
            self.fill_jpg_template(template_path, user_data, filled_jpg)
            
            # Step 3: Convert to PDF for compatibility
            self.convert_to_pdf(filled_jpg, filled_pdf)
            
            # Step 4: Create compressed upload versions
            self.compress_image(filled_jpg, compressed_jpg)
            self.convert_to_pdf(compressed_jpg, compressed_pdf)
            self.compress_pdf(compressed_pdf, compressed_pdf)
            
            # Step 5: Upload to WordPress (skip in test mode)
            if is_test_mode:
                wp_url_result = f"file://{os.path.abspath(compressed_pdf)}"
            else:
                wp_url_result = self.upload_to_wordpress(
                    compressed_pdf,
                    wp_url,
                    username,
                    app_password,
                    upload_filename=wordpress_upload_name
                )
            
            # Step 6: Generate QR code
            self.generate_qr_code(wp_url_result, qr_path)
            
            # Step 7: Embed QR in high-quality image
            self.embed_qr_in_image(filled_jpg, qr_path, final_jpg)
            
            # Step 8: Convert final image to PDF
            self.convert_to_pdf(final_jpg, final_pdf)

            result_data = {
                'final_jpg': final_jpg,
                'final_pdf': final_pdf,
                'wordpress_url': wp_url_result,
                'qr_target_link': wp_url_result,
                'qr_code': qr_path,
                'compressed_jpg': compressed_jpg,
                'compressed_pdf': compressed_pdf,
                'filled_jpg': filled_jpg,
                'filled_pdf': filled_pdf,
                'test_mode': is_test_mode,
                'folder_name': folder_name
            }

            # Step 9: Update license register
            if register_path is None:
                register_path = "license_register.xlsx"
            register_path = self.update_license_register(user_data, result_data, register_path)
            
            return result_data
            
        except Exception as e:
            raise


def main():
    """Main function to run the PIL-based license generator."""
    try:
        import qrcode
        from PIL import Image, ImageDraw, ImageFont
    except ImportError as e:
        return
    
    # Load WordPress credentials from environment
    wp_url = os.getenv('WP_URL')
    username = os.getenv('WP_USERNAME')
    app_password = os.getenv('WP_APP_PASSWORD')
    
    if not wp_url:
        return
    if not username:
        return
    if not app_password:
        return
    
    # Template JPG path
    template_path = "CPPL-Licence-Blank_page-0001.jpg"
    
    if not os.path.exists(template_path):
        return
    
    # Initialize and run the generator
    generator = LicenseGeneratorPIL()
    generator.run_complete_pipeline(template_path, wp_url, username, app_password)


if __name__ == "__main__":
    main()