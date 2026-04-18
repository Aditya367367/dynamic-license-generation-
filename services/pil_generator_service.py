from __future__ import annotations

import base64
import importlib.util
import io
import mimetypes
import os
import sys
import tempfile
import zipfile
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook


FIELD_NAME_MAP = {
    "licence_number": "Licence Number",
    "validity_from": "Validity From",
    "name_of_licensee": "Name of the Licensee",
    "type_of_premise": "Type of Premise",
    "license_category": "License Category",
    "address_of_premise": "Address of Premise",
}

FILE_LABELS = {
    "filled_jpg": "Filled JPG",
    "filled_pdf": "Filled PDF",
    "compressed_jpg": "Compressed JPG",
    "compressed_pdf": "Compressed PDF",
    "qr_code": "QR Code",
    "final_jpg": "Final JPG",
    "final_pdf": "Final PDF",
    "license_register": "Excel Register",
}


class LicensePILGeneratorService:
    """Bridge the standalone PIL generator into the Django app."""

    def __init__(self) -> None:
        self.source_dir = Path(settings.LICENSE_GENERATOR_SOURCE_DIR).resolve()
        self.output_dir = Path(settings.LICENSE_GENERATOR_OUTPUT_DIR).resolve()
        self.template_path = Path(settings.LICENSE_GENERATOR_TEMPLATE_PATH).resolve()
        self.field_mapping_path = Path(settings.LICENSE_GENERATOR_FIELD_MAPPING_PATH).resolve()
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._module = self._load_generator_module()

    def _load_generator_module(self):
        module_path = Path(__file__).parent / "license_generator_pil.py"
        spec = importlib.util.spec_from_file_location("license_generator_pil_bridge", str(module_path))
        if spec is None or spec.loader is None:
            raise RuntimeError("Unable to load license generator module.")

        module = importlib.util.module_from_spec(spec)
        sys.modules[spec.name] = module
        spec.loader.exec_module(module)
        return module

    def _build_url(self, request, relative_path: str, download: bool = False) -> str:
        url = request.build_absolute_uri(f"/api/licenses/files/{relative_path}")
        return f"{url}?download=1" if download else url

    def _relative_output_path(self, absolute_path: str) -> str:
        resolved = Path(absolute_path).resolve()
        return resolved.relative_to(self.output_dir).as_posix()

    def _sanitize_filename(self, text: str) -> str:
        """Sanitize text for use in filename."""
        import re
        sanitized = re.sub(r'[^\w\s-]', '', text)
        sanitized = re.sub(r'[-\s]+', '_', sanitized)
        return sanitized.strip('_')[:50]

    def _build_register_entry(self, request, path: str, label: str) -> dict:
        resolved = Path(path).resolve()
        relative_path = self._relative_output_path(str(resolved))
        return {
            "label": label,
            "name": resolved.name,
            "preview_url": self._build_url(request, relative_path),
            "download_url": self._build_url(request, relative_path, download=True),
        }

    def _build_zip_bundle_entry(self, result: dict) -> dict | None:
        folder_name = result.get("folder_name")
        if not folder_name:
            return None

        folder_path = Path(folder_name).resolve()
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for key in FILE_LABELS:
                if key == "license_register":
                    continue
                file_path = result.get(key)
                if file_path and Path(file_path).exists():
                    resolved = Path(file_path).resolve()
                    archive.write(resolved, arcname=f"{folder_path.name}/{resolved.name}")

        zip_data = base64.b64encode(zip_buffer.getvalue()).decode("utf-8")
        return {
            "label": "Download Folder",
            "name": f"{folder_path.name}.zip",
            "preview_kind": None,
            "mime_type": "application/zip",
            "data_url": f"data:application/zip;base64,{zip_data}",
        }

    def _build_memory_file_entry(self, path: str, label: str) -> dict:
        resolved = Path(path).resolve()
        suffix = resolved.suffix.lower()
        preview_kind = None
        if suffix in {".jpg", ".jpeg", ".png"}:
            preview_kind = "image"
        elif suffix == ".pdf":
            preview_kind = "pdf"

        mime_type = mimetypes.guess_type(resolved.name)[0] or "application/octet-stream"
        file_data = base64.b64encode(resolved.read_bytes()).decode("utf-8")
        data_url = f"data:{mime_type};base64,{file_data}"

        return {
            "label": label,
            "name": resolved.name,
            "preview_kind": preview_kind,
            "mime_type": mime_type,
            "data_url": data_url,
        }

    def _build_result_payload(self, request, result: dict) -> dict:
        files = []
        for key, label in FILE_LABELS.items():
            file_path = result.get(key)
            if key == "license_register":
                continue
            if file_path and Path(file_path).exists():
                files.append(self._build_memory_file_entry(file_path, label))

        return {
            "qr_target_link": result.get("qr_target_link") or result.get("wordpress_url"),
            "license_register": self._build_register_entry(request, result["license_register"], "Excel Register")
            if result.get("license_register")
            else None,
            "files": files,
            "download_bundle": self._build_zip_bundle_entry(result),
        }

    def generate(self, request, payload: dict) -> dict:
        register_path = self.output_dir / "license_register.xlsx"
        user_data = {display_name: payload[field_name] for field_name, display_name in FIELD_NAME_MAP.items()}

        # Create a subdirectory for this license generation
        license_number = user_data.get('Licence Number', 'UNKNOWN')
        licensee_name = user_data.get('Name of the Licensee', 'UNKNOWN')
        folder_name = f"{self._sanitize_filename(license_number)}_{self._sanitize_filename(licensee_name)}"
        license_output_dir = self.output_dir / folder_name
        license_output_dir.mkdir(parents=True, exist_ok=True)
        
        generator = self._module.LicenseGeneratorPIL(
            config_path=str(self.field_mapping_path),
        )
        result = generator.run_complete_pipeline(
            template_path=str(self.template_path),
            wp_url=os.getenv("WP_URL", ""),
            username=os.getenv("WP_USERNAME", ""),
            app_password=os.getenv("WP_APP_PASSWORD", ""),
            user_data=user_data,
            register_path=str(register_path),
            output_dir=str(license_output_dir),
        )
        return self._build_result_payload(request, result)

    def get_records(self, limit: int = 20) -> list[dict]:
        register_path = self.output_dir / "license_register.xlsx"
        if not register_path.exists():
            return []

        workbook = load_workbook(register_path, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(value or "").strip() for value in rows[0]]
        records = []
        for row in rows[1:]:
            if not any(row):
                continue
            record = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            records.append(
                {
                    "generated_at": record.get("Generated At"),
                    "licence_number": record.get("Licence Number"),
                    "validity_from": record.get("Validity From"),
                    "name_of_licensee": record.get("Name of the Licensee"),
                    "type_of_premise": record.get("Type of Premise"),
                    "license_category": record.get("License Category"),
                    "address_of_premise": record.get("Address of Premise"),
                    "link": record.get("Link"),
                }
            )

        return list(reversed(records))[:limit]

    def get_register_file(self, request) -> dict | None:
        register_path = self.output_dir / "license_register.xlsx"
        if not register_path.exists():
            return None
        return self._build_register_entry(request, str(register_path), "Excel Register")
